from openpyxl import load_workbook
import pymysql

db = pymysql.connect(host='127.0.0.1', port=3306, user='root', db='civil', password='123.com', charset='utf8')
cursor = db.cursor()

data = load_workbook('2033874.xlsx')

sheet = data.get_sheet_names()[0]

ws = data[sheet]

for params in ws.iter_rows(min_row=4):
    position_level = params[0].value
    job_category = params[1].value
    department_name = params[2].value
    position_code = params[4].value
    job_title = params[5].value
    job_profile = params[6].value
    num = params[7].value
    outlook = params[8].value
    major = params[9].value
    education = params[10].value
    age = params[11].value
    other_conditions = params[12].value
    accounts = params[13].value
    test_subjects = params[14].value
    remarks = params[15].value

    cursor.execute(
        "insert into civil_servant (position_level, job_category, department_name, position_code, job_title, job_profile, num, outlook, major, education, age, other_conditions, accounts, test_subjects, remarks) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
        (position_level, job_category, department_name, position_code,
         job_title, job_profile, num, outlook, major, education, age,
         other_conditions, accounts, test_subjects, remarks))

db.commit()
db.close()
